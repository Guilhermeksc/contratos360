from rest_framework import viewsets, views, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from django.db.models import Q, Count, Sum, F
from django.http import HttpResponse
from django.core.exceptions import ValidationError
from decimal import Decimal
import io
import openpyxl
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)

from .models import AmparoLegal, Compra, ItemCompra, Modalidade, ModoDisputa, ResultadoItem, Fornecedor
from django_licitacao360.apps.uasgs.models import Uasg
from .serializers import (
    CompraSerializer,
    CompraDetalhadaSerializer,
    CompraListagemSerializer,
    UnidadePorAnoSerializer,
    AnosUnidadesComboSerializer,
    ItemCompraSerializer,
    ResultadoItemSerializer,
    FornecedorSerializer,
    ItemResultadoMergeSerializer,
    ModalidadeAgregadaSerializer,
    FornecedorAgregadoSerializer,
)


class FornecedorViewSet(viewsets.ModelViewSet):
    queryset = Fornecedor.objects.all()
    serializer_class = FornecedorSerializer
    permission_classes = [AllowAny]
    search_fields = ["cnpj_fornecedor", "razao_social"]


class CompraViewSet(viewsets.ModelViewSet):
    queryset = Compra.objects.all()
    serializer_class = CompraSerializer
    permission_classes = [AllowAny]
    filterset_fields = ["ano_compra", "codigo_unidade", "modalidade", "amparo_legal", "modo_disputa"]
    search_fields = ["numero_compra", "objeto_compra", "numero_processo"]
    
    @action(detail=False, methods=['get'], url_path='por-unidade/(?P<codigo_unidade>[^/.]+)', permission_classes=[AllowAny])
    def por_unidade(self, request, codigo_unidade=None):
        """Retorna compras filtradas por código_unidade"""
        try:
            if not codigo_unidade:
                return Response(
                    {'error': 'Código da unidade é obrigatório'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            compras = self.queryset.filter(codigo_unidade=codigo_unidade).order_by('-ano_compra', '-sequencial_compra')
            serializer = self.get_serializer(compras, many=True)
            return Response(serializer.data)
        except Exception as e:
            logger.error(f"Erro ao buscar compras por unidade {codigo_unidade}: {str(e)}")
            return Response(
                {'error': 'Erro ao processar requisição'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'], url_path='modalidades-agregadas/(?P<codigo_unidade>[^/.]+)', permission_classes=[AllowAny])
    def modalidades_agregadas(self, request, codigo_unidade=None):
        """Retorna modalidades agregadas por código_unidade"""
        try:
            if not codigo_unidade:
                return Response(
                    {'error': 'Código da unidade é obrigatório'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            modalidades = (
                Compra.objects
                .filter(codigo_unidade=codigo_unidade)
                .select_related('modalidade')
                .values('ano_compra', 'modalidade_id')
                .annotate(
                    quantidade_compras=Count('compra_id'),
                    valor_total_homologado=Sum('valor_total_homologado')
                )
                .order_by('ano_compra', 'modalidade_id')
            )
            # Formata para o serializer - busca objetos Modalidade completos
            modalidades_formatted = []
            for mod in modalidades:
                modalidade_obj = None
                modalidade_id = mod.get('modalidade_id')
                if modalidade_id:
                    try:
                        modalidade_obj = Modalidade.objects.get(pk=modalidade_id)
                    except Modalidade.DoesNotExist:
                        pass
                
                modalidades_formatted.append({
                    'ano_compra': mod['ano_compra'],
                    'modalidade_id': modalidade_id,
                    'modalidade': modalidade_obj,
                    'quantidade_compras': mod['quantidade_compras'],
                    'valor_total_homologado': mod['valor_total_homologado'],
                })
            modalidades = modalidades_formatted
            serializer = ModalidadeAgregadaSerializer(modalidades, many=True)
            return Response(serializer.data)
        except Exception as e:
            logger.error(f"Erro ao buscar modalidades agregadas para unidade {codigo_unidade}: {str(e)}")
            return Response(
                {'error': 'Erro ao processar requisição'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'], url_path='itens-resultado-merge/(?P<codigo_unidade>[^/.]+)', permission_classes=[AllowAny])
    def itens_resultado_merge(self, request, codigo_unidade=None):
        """Retorna merge de itens com resultados por código_unidade"""
        try:
            if not codigo_unidade:
                return Response(
                    {'error': 'Código da unidade é obrigatório'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            compras = Compra.objects.filter(codigo_unidade=codigo_unidade)
            compra_ids = list(compras.values_list('compra_id', flat=True))
            
            # Otimização: usar prefetch_related para evitar N+1 queries
            itens = ItemCompra.objects.filter(
                compra_id__in=compra_ids
            ).select_related('compra').prefetch_related('resultados__fornecedor')
            
            result_data = []
            for item in itens:
                resultado = item.resultados.first()  # Pega o primeiro resultado se existir
                
                # Calcula percentual de desconto
                percentual_desconto = None
                if item.valor_total_estimado and resultado and resultado.valor_total_homologado:
                    if item.valor_total_estimado > 0:
                        percentual_desconto = (
                            (item.valor_total_estimado - resultado.valor_total_homologado) 
                            / item.valor_total_estimado
                        ) * 100
                
                result_data.append({
                    'ano_compra': item.compra.ano_compra,
                    'sequencial_compra': item.compra.sequencial_compra,
                    'numero_item': item.numero_item,
                    'descricao': item.descricao,
                    'unidade_medida': item.unidade_medida,
                    'valor_unitario_estimado': item.valor_unitario_estimado,
                    'valor_total_estimado': item.valor_total_estimado,
                    'quantidade': item.quantidade,
                    'situacao_compra_item_nome': item.situacao_compra_item_nome,
                    'cnpj_fornecedor': resultado.fornecedor.cnpj_fornecedor if resultado else None,
                    'valor_total_homologado': resultado.valor_total_homologado if resultado else None,
                    'valor_unitario_homologado': resultado.valor_unitario_homologado if resultado else None,
                    'quantidade_homologada': resultado.quantidade_homologada if resultado else None,
                    'percentual_desconto': percentual_desconto,
                    'razao_social': resultado.fornecedor.razao_social if resultado else None,
                })
            
            serializer = ItemResultadoMergeSerializer(result_data, many=True)
            return Response(serializer.data)
        except Exception as e:
            logger.error(f"Erro ao buscar itens-resultado merge para unidade {codigo_unidade}: {str(e)}")
            return Response(
                {'error': 'Erro ao processar requisição'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'], url_path='fornecedores-agregados/(?P<codigo_unidade>[^/.]+)', permission_classes=[AllowAny])
    def fornecedores_agregados(self, request, codigo_unidade=None):
        """Retorna fornecedores agregados por código_unidade"""
        try:
            if not codigo_unidade:
                return Response(
                    {'error': 'Código da unidade é obrigatório'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            compras = Compra.objects.filter(codigo_unidade=codigo_unidade)
            compra_ids = list(compras.values_list('compra_id', flat=True))
            
            # Otimização: usar select_related e valores diretamente
            fornecedores = (
                ResultadoItem.objects
                .filter(item_compra__compra_id__in=compra_ids)
                .select_related('fornecedor', 'item_compra')
                .values('fornecedor__cnpj_fornecedor', 'fornecedor__razao_social')
                .annotate(
                    valor_total_homologado=Sum('valor_total_homologado')
                )
                .order_by('fornecedor__cnpj_fornecedor')
            )
            
            result_data = []
            for f in fornecedores:
                result_data.append({
                    'cnpj_fornecedor': f['fornecedor__cnpj_fornecedor'],
                    'razao_social': f['fornecedor__razao_social'],
                    'valor_total_homologado': f['valor_total_homologado'] or Decimal('0'),
                })
            
            serializer = FornecedorAgregadoSerializer(result_data, many=True)
            return Response(serializer.data)
        except Exception as e:
            logger.error(f"Erro ao buscar fornecedores agregados para unidade {codigo_unidade}: {str(e)}")
            return Response(
                {'error': 'Erro ao processar requisição'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        except Exception as e:
            logger.error(f"Erro ao buscar fornecedores agregados para unidade {codigo_unidade}: {str(e)}")
            return Response(
                {'error': 'Erro ao processar requisição'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'], url_path='itens-por-modalidade/(?P<codigo_unidade>[^/.]+)', permission_classes=[AllowAny])
    def itens_por_modalidade(self, request, codigo_unidade=None):
        """Retorna itens agrupados por modalidade"""
        try:
            if not codigo_unidade:
                return Response(
                    {'error': 'Código da unidade é obrigatório'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            modalidade_id = request.query_params.get('modalidade_id')
            
            compras = Compra.objects.filter(codigo_unidade=codigo_unidade)
            if modalidade_id:
                compras = compras.filter(modalidade_id=modalidade_id)
            
            compra_ids = list(compras.values_list('compra_id', flat=True))
            itens = ItemCompra.objects.filter(compra_id__in=compra_ids).select_related('compra').prefetch_related('resultados__fornecedor')
            
            result_data = []
            for item in itens:
                resultado = item.resultados.first()
                
                # Calcula percentual de desconto
                percentual_desconto = None
                if item.valor_total_estimado and resultado and resultado.valor_total_homologado:
                    if item.valor_total_estimado > 0:
                        percentual_desconto = (
                            (item.valor_total_estimado - resultado.valor_total_homologado) 
                            / item.valor_total_estimado
                        ) * 100
                
                result_data.append({
                    'ano_compra': item.compra.ano_compra,
                    'sequencial_compra': item.compra.sequencial_compra,
                    'numero_item': item.numero_item,
                    'cnpj_fornecedor': resultado.fornecedor.cnpj_fornecedor if resultado else None,
                    'razao_social': resultado.fornecedor.razao_social if resultado else None,
                    'descricao': item.descricao,
                    'quantidade': item.quantidade,
                    'valor_total_estimado': item.valor_total_estimado,
                    'valor_total_homologado': resultado.valor_total_homologado if resultado else None,
                    'percentual_desconto': percentual_desconto,
                })
            
            serializer = ItemResultadoMergeSerializer(result_data, many=True)
            return Response(serializer.data)
        except Exception as e:
            logger.error(f"Erro ao buscar itens por modalidade para unidade {codigo_unidade}: {str(e)}")
            return Response(
                {'error': 'Erro ao processar requisição'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'], url_path='export-xlsx/(?P<codigo_unidade>[^/.]+)', permission_classes=[AllowAny])
    def export_xlsx(self, request, codigo_unidade=None):
        """Exporta dados para XLSX equivalente ao script Python"""
        try:
            if not codigo_unidade:
                return Response(
                    {'error': 'Código da unidade é obrigatório'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            from django.db.models import Prefetch
            
            # --- Compras ---
            compras = Compra.objects.filter(codigo_unidade=codigo_unidade)
            
            if not compras.exists():
                return Response(
                    {'error': f'Nenhuma compra encontrada para o código de unidade {codigo_unidade}'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            compra_ids = list(compras.values_list('compra_id', flat=True))
            
            # --- Itens ---
            itens = ItemCompra.objects.filter(
                compra_id__in=compra_ids
            ).select_related('compra').prefetch_related('resultados__fornecedor')
            
            # --- Resultados ---
            resultados_dict = {}
            for item in itens:
                resultado = item.resultados.first()
                if resultado:
                    resultados_dict[item.item_id] = resultado
            
            # --- Fornecedores ---
            # Otimização: carregar apenas fornecedores relacionados aos resultados
            fornecedores_cnpjs = set()
            for item in itens:
                resultado = item.resultados.first()
                if resultado:
                    fornecedores_cnpjs.add(resultado.fornecedor.cnpj_fornecedor)
            
            fornecedores_dict = {
                f.cnpj_fornecedor: f.razao_social 
                for f in Fornecedor.objects.filter(cnpj_fornecedor__in=fornecedores_cnpjs)
            }
            
            # --- Merge Itens + Resultados ---
            merge_data = []
            for item in itens:
                resultado = resultados_dict.get(item.item_id)
                
                percentual_desconto = None
                if item.valor_total_estimado and resultado and resultado.valor_total_homologado:
                    if item.valor_total_estimado > 0:
                        percentual_desconto = (
                            (item.valor_total_estimado - resultado.valor_total_homologado) 
                            / item.valor_total_estimado
                        ) * 100
                
                merge_data.append({
                    'ano_compra': item.compra.ano_compra,
                    'sequencial_compra': item.compra.sequencial_compra,
                    'numero_item': item.numero_item,
                    'descricao': item.descricao,
                    'unidade_medida': item.unidade_medida,
                    'valor_unitario_estimado': item.valor_unitario_estimado,
                    'valor_total_estimado': item.valor_total_estimado,
                    'quantidade': item.quantidade,
                    'situacao_compra_item_nome': item.situacao_compra_item_nome,
                    'cnpj_fornecedor': resultado.fornecedor.cnpj_fornecedor if resultado else None,
                    'valor_total_homologado': resultado.valor_total_homologado if resultado else None,
                    'valor_unitario_homologado': resultado.valor_unitario_homologado if resultado else None,
                    'quantidade_homologada': resultado.quantidade_homologada if resultado else None,
                    'percentual_desconto': percentual_desconto,
                })
            
            # --- Modalidades Agregadas ---
            modalidades = (
                compras.select_related('modalidade')
                .values('ano_compra', 'modalidade_id')
                .annotate(
                    quantidade_compras=Count('compra_id'),
                    valor_total_homologado=Sum('valor_total_homologado')
                )
            )
            # Formata para o serializer - busca objetos Modalidade completos
            modalidades_formatted = []
            for mod in modalidades:
                modalidade_obj = None
                modalidade_id = mod.get('modalidade_id')
                if modalidade_id:
                    try:
                        modalidade_obj = Modalidade.objects.get(pk=modalidade_id)
                    except Modalidade.DoesNotExist:
                        pass
                
                modalidades_formatted.append({
                    'ano_compra': mod['ano_compra'],
                    'modalidade_id': modalidade_id,
                    'modalidade': modalidade_obj,
                    'quantidade_compras': mod['quantidade_compras'],
                    'valor_total_homologado': mod['valor_total_homologado'],
                })
            modalidades = modalidades_formatted
            
            # --- Fornecedores Agregados ---
            fornecedores_agregados = []
            fornecedores_valores = {}
            for item in merge_data:
                if item['cnpj_fornecedor'] and item['valor_total_homologado']:
                    cnpj = item['cnpj_fornecedor']
                    if cnpj not in fornecedores_valores:
                        fornecedores_valores[cnpj] = Decimal('0')
                    fornecedores_valores[cnpj] += item['valor_total_homologado']
            
            for cnpj, valor in fornecedores_valores.items():
                fornecedores_agregados.append({
                    'cnpj_fornecedor': cnpj,
                    'razao_social': fornecedores_dict.get(cnpj),
                    'valor_total_homologado': valor,
                })
            
            # --- Preparar dados das compras ---
            compras_data = []
            for compra in compras:
                compra_dict = CompraSerializer(compra).data
                compras_data.append(compra_dict)
            
            # --- Adicionar razao_social ao merge ---
            for item in merge_data:
                item['razao_social'] = fornecedores_dict.get(item['cnpj_fornecedor']) if item['cnpj_fornecedor'] else None
            
            # --- Criar dicionário de compras para lookup rápido ---
            compras_lookup = {}
            for compra in compras:
                key = (compra.ano_compra, compra.sequencial_compra)
                compras_lookup[key] = compra.modalidade if compra.modalidade else None
            
            # --- Itens por Modalidade ---
            modalidades_unicas = compras.select_related('modalidade').values_list('modalidade', flat=True).distinct()
            
            # --- Criar XLSX ---
            output = io.BytesIO()
            wb = openpyxl.Workbook()
            
            # Remover sheet padrão
            wb.remove(wb.active)
            
            # Função auxiliar para nome de sheet
            def _sheet_name_from_modalidade(modalidade):
                if not modalidade:
                    return "modalidade"
                nome = str(modalidade).strip()
                nome = nome.replace("/", "-").replace("\\", "-").replace(":", "-")
                nome = nome.replace("*", " ").replace("?", " ").replace("[", "(").replace("]", ")")
                nome = " ".join(nome.split())
                return nome[:31] if nome else "modalidade"
            
            # Sheet: compras
            ws_compras = wb.create_sheet("compras")
            if compras_data:
                headers = list(compras_data[0].keys())
                ws_compras.append(headers)
                for compra in compras_data:
                    ws_compras.append([compra.get(h) for h in headers])
            
            # Sheet: itens_resultado_merge
            ws_merge = wb.create_sheet("itens_resultado_merge")
            if merge_data:
                headers = list(merge_data[0].keys())
                ws_merge.append(headers)
                for item in merge_data:
                    ws_merge.append([item.get(h) for h in headers])
            
            # Sheet: modalidades
            ws_modalidades = wb.create_sheet("modalidades")
            ws_modalidades.append(['ano_compra', 'modalidade_id', 'modalidade_nome', 'quantidade_compras', 'valor_total_homologado'])
            for mod in modalidades:
                modalidade_nome = mod.get('modalidade').nome if mod.get('modalidade') else None
                ws_modalidades.append([
                    mod['ano_compra'],
                    mod.get('modalidade_id'),
                    modalidade_nome,
                    mod['quantidade_compras'],
                    mod['valor_total_homologado']
                ])
            
            # Sheet: fornecedores
            ws_fornecedores = wb.create_sheet("fornecedores")
            ws_fornecedores.append(['cnpj_fornecedor', 'razao_social', 'valor_total_homologado'])
            for f in fornecedores_agregados:
                ws_fornecedores.append([
                    f['cnpj_fornecedor'],
                    f['razao_social'],
                    f['valor_total_homologado']
                ])
            
            # Sheet: inexigibilidade
            inexigibilidade_modalidade = Modalidade.objects.filter(nome='Inexigibilidade').first()
            inexigibilidade_data = [
                item for item in merge_data
                if compras_lookup.get((item['ano_compra'], item['sequencial_compra'])) == inexigibilidade_modalidade
            ]
            ws_inex = wb.create_sheet("inexigibilidade")
            if inexigibilidade_data:
                cols_inex = ['ano_compra', 'sequencial_compra', 'numero_item', 'cnpj_fornecedor', 
                            'razao_social', 'descricao', 'quantidade', 'valor_total_estimado', 
                            'valor_total_homologado', 'percentual_desconto']
                ws_inex.append(cols_inex)
                for item in inexigibilidade_data:
                    ws_inex.append([item.get(c) for c in cols_inex])
            
            # Sheets por modalidade
            for modalidade_id in modalidades_unicas:
                if not modalidade_id:
                    continue
                
                try:
                    modalidade_obj = Modalidade.objects.get(id=modalidade_id)
                except Modalidade.DoesNotExist:
                    continue
                
                # Pula inexigibilidade (já tem sheet própria)
                if modalidade_obj.nome.lower() == 'inexigibilidade':
                    continue
                
                modalidade_data = [
                    item for item in merge_data
                    if compras_lookup.get((item['ano_compra'], item['sequencial_compra'])) == modalidade_obj
                ]
                
                if not modalidade_data:
                    continue
                
                sheet_name = _sheet_name_from_modalidade(modalidade_obj.nome)
                ws_mod = wb.create_sheet(sheet_name)
                cols_mod = ['ano_compra', 'sequencial_compra', 'numero_item', 'cnpj_fornecedor', 
                           'razao_social', 'descricao', 'quantidade', 'valor_total_estimado', 
                           'valor_total_homologado', 'percentual_desconto']
                ws_mod.append(cols_mod)
                for item in modalidade_data:
                    ws_mod.append([item.get(c) for c in cols_mod])
            
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="compras_{codigo_unidade}.xlsx"'
            return response
        except Exception as e:
            logger.error(f"Erro ao exportar XLSX para unidade {codigo_unidade}: {str(e)}")
            return Response(
                {'error': 'Erro ao gerar arquivo XLSX'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ItemCompraViewSet(viewsets.ModelViewSet):
    queryset = ItemCompra.objects.all()
    serializer_class = ItemCompraSerializer
    permission_classes = [AllowAny]
    filterset_fields = ["compra", "tem_resultado", "situacao_compra_item_nome"]
    search_fields = ["descricao"]


class ResultadoItemViewSet(viewsets.ModelViewSet):
    queryset = ResultadoItem.objects.all()
    serializer_class = ResultadoItemSerializer
    permission_classes = [AllowAny]
    filterset_fields = ["item_compra", "fornecedor", "status"]


class CompraDetalhadaView(views.APIView):
    """Endpoint para buscar compra detalhada por codigo_unidade, numero_compra, ano_compra e modalidade"""
    permission_classes = [AllowAny]

    def get(self, request):
        codigo_unidade = request.query_params.get('codigo_unidade')
        numero_compra = request.query_params.get('numero_compra')
        ano_compra = request.query_params.get('ano_compra')
        modalidade = request.query_params.get('modalidade')

        # Validação dos parâmetros obrigatórios
        if not codigo_unidade:
            return Response(
                {'error': 'Parâmetro codigo_unidade é obrigatório'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not numero_compra:
            return Response(
                {'error': 'Parâmetro numero_compra é obrigatório'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not ano_compra:
            return Response(
                {'error': 'Parâmetro ano_compra é obrigatório'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not modalidade:
            return Response(
                {'error': 'Parâmetro modalidade é obrigatório'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            ano_compra = int(ano_compra)
            modalidade_id = int(modalidade)
        except ValueError:
            return Response(
                {'error': 'ano_compra e modalidade devem ser números inteiros'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Buscar a compra com os parâmetros fornecidos
            compra = Compra.objects.filter(
                codigo_unidade=codigo_unidade,
                numero_compra=numero_compra,
                ano_compra=ano_compra,
                modalidade_id=modalidade_id
            ).select_related(
                'modalidade',
                'amparo_legal',
                'modo_disputa'
            ).prefetch_related(
                'itens__resultados__fornecedor'
            ).first()

            if not compra:
                return Response(
                    {'error': 'Compra não encontrada com os parâmetros fornecidos'},
                    status=status.HTTP_404_NOT_FOUND
                )

            serializer = CompraDetalhadaSerializer(compra)
            return Response(serializer.data)

        except Exception as e:
            logger.error(f"Erro ao buscar compra detalhada: {str(e)}")
            return Response(
                {'error': 'Erro ao processar requisição'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class CompraListagemView(views.APIView):
    """Endpoint para listar compras por codigo_unidade e ano_compra"""
    permission_classes = [AllowAny]

    def get(self, request):
        codigo_unidade = request.query_params.get('codigo_unidade')
        ano_compra = request.query_params.get('ano_compra')

        # Validação dos parâmetros obrigatórios
        if not codigo_unidade:
            return Response(
                {'error': 'Parâmetro codigo_unidade é obrigatório'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not ano_compra:
            return Response(
                {'error': 'Parâmetro ano_compra é obrigatório'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            ano_compra = int(ano_compra)
        except ValueError:
            return Response(
                {'error': 'ano_compra deve ser um número inteiro'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Buscar todas as compras com os parâmetros fornecidos
            compras = Compra.objects.filter(
                codigo_unidade=codigo_unidade,
                ano_compra=ano_compra
            ).select_related(
                'modalidade',
                'amparo_legal',
                'modo_disputa'
            ).order_by('-sequencial_compra')

            serializer = CompraListagemSerializer(compras, many=True)
            return Response({
                'count': compras.count(),
                'results': serializer.data
            })

        except Exception as e:
            logger.error(f"Erro ao listar compras: {str(e)}")
            return Response(
                {'error': 'Erro ao processar requisição'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class UnidadesPorAnoView(views.APIView):
    """Endpoint para listar todos os codigo_unidade por ano_compra"""
    permission_classes = [AllowAny]

    def get(self, request):
        ano_compra = request.query_params.get('ano_compra')

        # Validação do parâmetro obrigatório
        if not ano_compra:
            return Response(
                {'error': 'Parâmetro ano_compra é obrigatório'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            ano_compra = int(ano_compra)
        except ValueError:
            return Response(
                {'error': 'ano_compra deve ser um número inteiro'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Buscar todas as unidades distintas com compras no ano especificado
            unidades = (
                Compra.objects
                .filter(ano_compra=ano_compra)
                .values('codigo_unidade', 'ano_compra')
                .annotate(
                    quantidade_compras=Count('compra_id'),
                    valor_total_estimado=Sum('valor_total_estimado'),
                    valor_total_homologado=Sum('valor_total_homologado')
                )
                .order_by('codigo_unidade')
            )

            serializer = UnidadePorAnoSerializer(unidades, many=True)
            return Response({
                'ano_compra': ano_compra,
                'count': len(unidades),
                'results': serializer.data
            })

        except Exception as e:
            logger.error(f"Erro ao listar unidades por ano: {str(e)}")
            return Response(
                {'error': 'Erro ao processar requisição'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class AnosUnidadesComboView(views.APIView):
    """Endpoint para retornar todos os anos disponíveis e códigos de unidade com sigla_om por ano (para combobox)"""
    permission_classes = [AllowAny]

    def get(self, request):
        try:
            # Buscar todos os anos distintos ordenados de forma decrescente
            anos_distintos = (
                Compra.objects
                .values_list('ano_compra', flat=True)
                .distinct()
                .order_by('-ano_compra')
            )
            anos_list = list(anos_distintos)

            # Buscar todos os códigos de unidade únicos de uma vez para otimizar
            todos_codigos_unidade = (
                Compra.objects
                .values_list('codigo_unidade', flat=True)
                .distinct()
            )
            
            # Criar um dicionário de código_unidade -> sigla_om para lookup rápido
            siglas_cache = {}
            codigos_para_buscar = []
            for codigo in todos_codigos_unidade:
                try:
                    codigo_int = int(codigo)
                    codigos_para_buscar.append(codigo_int)
                except (ValueError, TypeError):
                    pass
            
            # Buscar todas as UASGs de uma vez
            if codigos_para_buscar:
                uasgs = Uasg.objects.filter(uasg__in=codigos_para_buscar).values('uasg', 'sigla_om')
                for uasg in uasgs:
                    siglas_cache[str(uasg['uasg'])] = uasg['sigla_om']

            # Para cada ano, buscar os códigos de unidade distintos com sigla_om
            unidades_por_ano = {}
            for ano in anos_list:
                # Buscar códigos de unidade distintos para o ano
                codigos_unidade = (
                    Compra.objects
                    .filter(ano_compra=ano)
                    .values_list('codigo_unidade', flat=True)
                    .distinct()
                    .order_by('codigo_unidade')
                )
                
                # Criar lista de unidades com sigla_om usando o cache
                unidades_com_sigla = []
                for codigo in codigos_unidade:
                    sigla_om = siglas_cache.get(codigo)
                    unidades_com_sigla.append({
                        'codigo_unidade': codigo,
                        'sigla_om': sigla_om
                    })
                
                unidades_por_ano[str(ano)] = unidades_com_sigla

            return Response({
                'anos': anos_list,
                'unidades_por_ano': unidades_por_ano
            })

        except Exception as e:
            logger.error(f"Erro ao buscar anos e unidades para combobox: {str(e)}")
            return Response(
                {'error': 'Erro ao processar requisição'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
